import logging
from typing import Callable, List, Union, Set, Tuple, Any, Optional
from pathlib import Path
from random import choice
import numpy as np
from enum import Enum
from copy import deepcopy
from types import FrameType
from fuzzingbook.Grammars import Grammar
from fuzzingbook.Parser import EarleyParser
from fuzzingbook.ProbabilisticGrammarFuzzer import (
    is_valid_probabilistic_grammar,
    ProbabilisticGrammarMiner,
    ProbabilisticGrammarFuzzer,
)
from isla.derivation_tree import DerivationTree
import openpyxl
from evogfuzz.tournament_selection import Tournament
from evogfuzz.fitness_functions import fitness_function_except
from evogfuzz.fitness_functions import fitness_function_memory
from evogfuzz.fitness_functions import fitness_function_coverage
from evogfuzz import helper
from evogfuzz.oracle import OracleResult
from evogfuzz.input import Input
import json
import sys
import time
import tracemalloc
import inspect
class GrammarType(Enum):
    MUTATED = "mutated"
    LEARNED = "learned"

    def __str__(self):
        return self.value

def traceit(frame: FrameType, event: str, arg: Any) -> Optional[Callable]:
        """Trace program execution. To be passed to sys.settrace()."""
        if event == 'line':
            global coverage
            function_name = frame.f_code.co_name
            lineno = frame.f_lineno
            if(function_name =="buggy_function_cov"):
                coverage.append(lineno)
        return traceit
    
class EvoGFuzz:

    
    def __init__(
        self,
        grammar: Grammar,
        prop: Callable[[Union[Input, str]], OracleResult],
        inputs: List[str],
        fitness_function: Callable[
            [Input], float
        ],
        fitnessType: str = "failure", 
        which_rerun: int = 1,
        with_mutations: bool = True,
        iterations: int = 10,
        working_dir: Path = None,
    ):
        self.which_rerun = which_rerun
        self.with_mutations = with_mutations
        self.fitnessType = fitnessType
        self.grammar = grammar
        self._prop: Callable[[Input], OracleResult] = prop
        self.working_dir = working_dir
        self._probabilistic_grammars: List[Tuple[Grammar, GrammarType, float]] = []
        self._iteration: int = 0
        self._max_iterations: int = iterations
        self._number_individuals: int = 100
        self._parameter_lambda: float = 2
        self._elitism_rate: int = 5
        self._tournament_size: int = 10
        self._tournament_number: int = 10
        self._all_inputs = None
        self._avg_prev_data = 0
        self.fitness_function: Union[
            Callable[
                [Input],
                float,
            ],
            None,
        ] = fitness_function
        self._probabilistic_grammar_miner = ProbabilisticGrammarMiner(
            EarleyParser(self.grammar)
        )

        self.inputs = set()
        for inp in inputs:
            self.inputs.add(
                Input(
                    DerivationTree.from_parse_tree(
                        next(EarleyParser(grammar).parse(inp))
                    )
                )
            )

        # Apply patch to fuzzingbook
        helper.patch()
    
    def setup(self):

        match self.fitnessType: #fitness_function
            case "cputime": #fitness_function_memory
                for inp in self.inputs:
                #    inp.oracle, inp.exec_feature= self._prop(inp)
                    st = time.process_time()
                    try:
                        self._prop(inp)
                        ct = time.process_time() - st
                        inp.oracle, inp.exec_feature = OracleResult.NO_BUG, round(ct*10000, 4)
                    except:
                        ct = time.process_time() - st
                        inp.oracle, inp.exec_feature = OracleResult.BUG, round(ct*10000, 4)

            case "wctime":
                for inp in self.inputs:
                #    inp.oracle, inp.exec_feature= self._prop(inp)
                    st = time.time()
                    try:
                        self._prop(inp)
                        ct = time.time() - st
                        inp.oracle, inp.exec_feature = OracleResult.NO_BUG, round(ct*10000, 4)
                    except:
                        ct = time.time() - st
                        inp.oracle, inp.exec_feature = OracleResult.BUG, round(ct*10000, 4)

            case "failure":
                for inp in self.inputs:
                    try:
                        self._prop(inp)
                        inp.oracle = OracleResult.NO_BUG
                    except:
                        inp.oracle = OracleResult.BUG
            
            case "exception":
                #for inp in self.inputs:
                #    inp.oracle, inp.error_message = self._prop(inp)

                for inp in self.inputs:
                    try:
                        self._prop(inp)
                        inp.oracle, inp.error_message = OracleResult.NO_BUG, "no_ex_triggered"
                    except Exception as e: 
                        inp.oracle, inp.error_message = OracleResult.BUG, sys.exc_info()[0]
            case "memory":
                 for inp in self.inputs:
                    try:
                        tracemalloc.start()
                        tracemalloc.reset_peak()
                        self._prop(inp)
                        size, peak= tracemalloc.get_traced_memory()
                        tracemalloc.stop
                        inp.oracle, inp.exec_feature = OracleResult.NO_BUG, peak
                    except Exception as e: 
                        inp.oracle, inp.exec_feature = OracleResult.BUG, None
            case "coverage":
                 for inp in self.inputs:
                    try:
                        global coverage
                        coverage = []
                        sys.settrace(traceit)  # Turn on
                        self._prop(inp)
                        sys.settrace(None) #turn off
                        code = inspect.getsource(self._prop)
                        lengthOfCoverage = len(coverage)
                        lengthOfCode = len(code.splitlines())
                        percentCovered = lengthOfCoverage / lengthOfCode
                        inp.oracle, inp.exec_feature = OracleResult.NO_BUG, percentCovered
                    except Exception as e: 
                        inp.oracle, inp.exec_feature = OracleResult.BUG, e
        
        probabilistic_grammar = self._learn_probabilistic_grammar(self.inputs)
        self._probabilistic_grammars.append(
            (deepcopy(probabilistic_grammar), GrammarType.LEARNED, -1)
        )
        initial_population = self._generate_input_files(probabilistic_grammar)

        return initial_population

    def _save_population(self, column_name, inputs):
            # Load the workbook or create a new one if it doesn't exist
            try:
                wb = openpyxl.load_workbook(f'{self.fitnessType}_without_mutation.xlsx')
            except FileNotFoundError:
                wb = openpyxl.Workbook()
            
            # Select the active worksheet or create a new one if it doesn't exist
            try:
                ws = wb[f'{self.which_rerun}. run']
            except KeyError:
                ws = wb.create_sheet(f'{self.which_rerun}. run')

            # Add the column name to the first row
            ws.cell(row=1, column=ws.max_column+1, value=str(column_name)+". loop")

            # Add the inputs to the column
            for i, input_value in enumerate(inputs):
                ws.cell(row=i+2, column=ws.max_column, value=str(input_value))

            # Save the workbook
            wb.save(f'{self.fitnessType}_without_mutation.xlsx')

    def optimize(self) -> Grammar:
        logging.info("Optimizing with EvoGFuzz")
        while self._do_more_iterations():
            logging.info("Starting to optimize probabilities")

        # Return best Grammar
        return Grammar
            
    def fuzz(self):
        logging.info("Fuzzing with EvoGFuzz")
        best_grammar_counter = 0
        new_population: Set[Input] = self.setup()
        while self._do_more_iterations():
            logging.info(f"Starting iteration {self._iteration}")
            new_population = self._loop(new_population)

            this_round_counter = self._check_best_round(new_population)
            if(best_grammar_counter<this_round_counter):
                best_grammar_counter = this_round_counter
                best_grammar = self._probabilistic_grammars[-1][0]
                iteration = self._iteration

            if(self._iteration %2 == 0):
                self._save_population(self._iteration, new_population)
            self._iteration = self._iteration + 1

        if(self._iteration %2 == 0):
                self._save_population(self._iteration, new_population)
        self._save_grammars_to_text(best_grammar, iteration)
        self._finalize()


    def _save_grammars_to_text(self, grammar, iteration):

        with open(f'{self.fitnessType}_grammars_without_mutation.txt', 'a') as f:
            f.write(f"Results from the {self.which_rerun}. rerun\n\n")
            f.write("First grammar:\n")
            json.dump(self._probabilistic_grammars[0][0], f)
            f.write('\n\n')
            f.write(f"Best grammar: {iteration}\n")
            json.dump(grammar, f)
            f.write('\n')
            f.write("Last grammar: \n")
            json.dump(self._probabilistic_grammars[-1][0], f)
            f.write('\n\n')


    def _check_best_round(self, inputs):
        counter = 0
        for input in inputs:
            inp = str(input)
            match self.fitnessType:
                case "memory":
                    if ( "mem" in inp):
                        counter = counter+1   
                case "cputime":
                    if ("cpu" in inp):                     
                        counter = counter+1
                case "wctime":
                    if ( "time" in inp):
                        counter = counter+1
                case "coverage":
                    if ("cov" in inp):
                        counter = counter+1
                case "failure":
                    if ("error" in inp or "exception" in inp):
                        counter = counter+1
                case "exception":
                    if ("exception" in inp):
                        counter = counter+1
        return counter
            
    def _loop(self, test_inputs: Set[Input]):
        # obtain labels, execute samples (Initial Step, Activity 5)
        match self.fitnessType:
            case "cputime":
                for inp in test_inputs:
                #    inp.oracle, inp.exec_feature= self._prop(inp)
                    st = time.process_time()
                    try:
                        self._prop(inp)
                        ct = time.process_time() - st
                        inp.oracle, inp.exec_feature = OracleResult.NO_BUG, round(ct*10000, 4)
                    except:
                        ct = time.process_time() - st
                        inp.oracle, inp.exec_feature = OracleResult.BUG, round(ct*10000, 4)

            case "wctime":
                for inp in test_inputs:
                #    inp.oracle, inp.exec_feature= self._prop(inp)
                    st = time.time()
                    try:
                        self._prop(inp)
                        ct = time.time() - st
                        inp.oracle, inp.exec_feature = OracleResult.NO_BUG, round(ct*10000, 4)
                    except:
                        ct = time.time() - st
                        inp.oracle, inp.exec_feature = OracleResult.BUG, round(ct*10000, 4)
            case "failure":
                for inp in test_inputs:
                    try:
                        self._prop(inp)
                        inp.oracle = OracleResult.NO_BUG
                    except Exception as e: 
                        inp.oracle = OracleResult.BUG
            case "exception":
                for inp in test_inputs:
                    try:
                        self._prop(inp)
                        inp.oracle, inp.error_message = OracleResult.NO_BUG, "no_ex_triggered"
                    except Exception as e: 
                        inp.oracle, inp.error_message = OracleResult.BUG, sys.exc_info()[0]
            case "memory":
                 for inp in test_inputs:
                    try:
                        tracemalloc.start()
                        tracemalloc.reset_peak()
                        self._prop(inp)
                        size, peak= tracemalloc.get_traced_memory()
                        tracemalloc.stop
                        inp.oracle, inp.exec_feature = OracleResult.NO_BUG, peak
                    except Exception as e: 
                        inp.oracle, inp.exec_feature = OracleResult.BUG, None
            case "coverage":
                 for inp in test_inputs:
                    try:
                        global coverage
                        coverage = []
                        sys.settrace(traceit)  # Turn on
                        self._prop(inp)
                        sys.settrace(None) #turn off
                        code = inspect.getsource(self._prop)
                        lengthOfCoverage = len(coverage)
                        lengthOfCode = len(code.splitlines())
                        percentCovered = lengthOfCoverage / lengthOfCode
                        inp.oracle, inp.exec_feature = OracleResult.NO_BUG, percentCovered
                    except Exception as e: 
                        inp.oracle, inp.exec_feature = OracleResult.BUG, e
        # determine fitness of individuals
        for inp in test_inputs:
            inp.fitness = self.fitness_function(inp)
        # select fittest individuals
        fittest_individuals: Set[Input] = self._select_fittest_individuals(test_inputs)

        # learn new probabilistic grammar
        probabilistic_grammar = self._learn_probabilistic_grammar(fittest_individuals)
        self._probabilistic_grammars.append(
            (deepcopy(probabilistic_grammar), GrammarType.LEARNED, -1)
        )
        #HERE
        # mutate grammar
        if(self.with_mutations):
             # generate new population
            mutated_grammar = self._mutate_grammar(probabilistic_grammar)
            self._probabilistic_grammars.append((mutated_grammar, GrammarType.MUTATED, -1))
            return self._generate_input_files(mutated_grammar)
        else:
            return self._generate_input_files(probabilistic_grammar)
       
        #return self._generate_input_files(_probabilistic_grammar)
       

    def _do_more_iterations(self):
        #hier timer gucken
        if -1 == self._max_iterations:
            return True
        if self._iteration >= self._max_iterations:
            logging.info("Terminate due to maximal iterations reached")
            return False
        return True

    def _generate_input_files(self, probabilistic_grammar):
        logging.info("Generating new Test Inputs")
        probabilistic_fuzzer = ProbabilisticGrammarFuzzer(probabilistic_grammar)
        new_test_inputs = set()
        for _ in range(self._number_individuals):
            new_test_inputs.add(Input(
                DerivationTree.from_parse_tree(probabilistic_fuzzer.fuzz_tree())
                )
            )
        return new_test_inputs

    def _safe_fitness_for_grammar(self, sum_fitness: float):
        grammar_tuple: Tuple[Grammar, GrammarType, float] = self._probabilistic_grammars.pop()
        grammar, grammar_type, _ = grammar_tuple
        self._probabilistic_grammars.append((grammar, grammar_type, sum_fitness))

    def _select_fittest_individuals(
        self, test_inputs: Set[Input]
    ) -> Set[Input]:

        fittest_individuals = Tournament(
            test_inputs, self._tournament_number
        ).select_fittest_individuals()

        sum_fitness = sum([inp.fitness for inp in fittest_individuals])
        logging.debug(
            f"Current probabilistic grammar achieved a combined fitness of: {sum_fitness}"
        )
        self._safe_fitness_for_grammar(sum_fitness)

        return fittest_individuals

    def _learn_probabilistic_grammar(self, test_inputs: Set[Input]):
        logging.info("Learning new Grammar")

        input_strings = list(str(inp.tree) for inp in test_inputs)

        probabilistic_grammar = (
            self._probabilistic_grammar_miner.mine_probabilistic_grammar(input_strings)
        )
        assert is_valid_probabilistic_grammar(
            probabilistic_grammar
        ), "Exit! Newly generated Grammar is not valid!"

        return probabilistic_grammar

    @staticmethod
    def _mutate_grammar(probabilistic_grammar):
        logging.info("Mutating new Grammar")

        mutated_grammar = deepcopy(probabilistic_grammar)

        # TODO if prev_avg < avg_fitness*1.025 -> Mutate

        # Only select production_rules with > 2 alternatives
        filtered = list(
            filter(lambda x: len(mutated_grammar[x]) > 1, list(mutated_grammar))
        )
        selected = choice(filtered)
        logging.info(f"Selected rule {selected} to be mutated.")
        new_probs = np.random.random(size=len(mutated_grammar[selected]))
        new_probs /= new_probs.sum()

        for count, child in enumerate(mutated_grammar[selected]):
            child[1]["prob"] = list(new_probs)[count]

        for rule in mutated_grammar:
            logging.info(rule.ljust(30) + str(mutated_grammar[rule]))

        return mutated_grammar

    def _finalize(self):
        logging.info("Exiting EvoGFuzz!")
        logging.info("Final Grammar:")

        final_grammar = self._get_latest_grammar()
        for rule in final_grammar:
            logging.info(rule.ljust(30) + str(final_grammar[rule]))

    def _get_latest_grammar(self):
        return self._probabilistic_grammars[-1][0]
